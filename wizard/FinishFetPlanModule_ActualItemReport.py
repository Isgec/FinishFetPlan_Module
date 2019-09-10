from odoo import models, fields
import io
import base64
import openpyxl
from datetime import timedelta
from os import path
from openpyxl.styles import PatternFill
from tempfile import TemporaryFile


class FinishFetPlanModule_ActualItemReport(models.TransientModel):
    _name = 'finishfetplanmodule.actualitemreport'
    from_dt = fields.Date('Select Date from which to Update Actual ', required=True)
    name = fields.Char('Finish Fet Excel Sheet')
    upload_file = fields.Binary(string="Upload Revised Actual")
    uploadedfilename = fields.Char('File Name', size=256, default='Select Revised Actual')
    readfromexcel = fields.Text('Final Step')
    remarks = fields.Text('Remarks')

    def upload_excel(self, data, context=None):
        # Generating of the excel file to be read by openpyxl
        fillGRINDING = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        fillGOUGING = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        fillWELDING = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        file = base64.decodestring(self.upload_file)
        excel_fileobj = TemporaryFile('wb+')
        excel_fileobj.write(file)
        excel_fileobj.seek(0)

        # Create workbook
        wb = openpyxl.load_workbook(excel_fileobj, data_only=True)
        wb.active = 0
        worksheet = wb.active

        header_obj = self.env['finishfetplanmodule.itemplanheadertable']
        header_ids = header_obj.search([(1, '=', 1)])
        self.readfromexcel = ''
        itempos = 15
        relativedate = 0
        my_max_col = 120
        for thisheader_ids in header_ids:
            self.readfromexcel = self.readfromexcel + '{ Items : ' + thisheader_ids.name + '}'
            relativedate = 0
            my_max_col = 120
            for thisitems_ids in thisheader_ids.actualitemplan_id:
                if thisitems_ids.date >= self.from_dt:
                    jobroutingid = thisitems_ids.jobrouting_id.id
                    plandate = thisitems_ids.date
                    thisitems_ids.unlink()  # Delete Record  from Item table With particular

            for i in range(4, my_max_col + 1, 3):
                # Reading for Shift A
                getdt = self.from_dt + timedelta(relativedate)
                getcol = worksheet.cell(row=itempos, column=i)
                jobrouting_obj = self.env['finishfetplanmodule.jobroutingtable']
                jobrouting_id = jobrouting_obj.search([('colour', '=', str(getcol.fill)[139:147])])
                for thisjob in jobrouting_id:
                    thisheader_ids.actualitemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id, 'jobrouting_id': thisjob.id,
                         'date': getdt,
                         'name': 'Added Shift A',
                         'shift_a': getcol.value,
                         'shift_b': 0,
                         'shift_c': 0})

                # Reading for Shift B
                getdt = self.from_dt + timedelta(relativedate)
                getcol = worksheet.cell(row=itempos, column=i + 1)
                jobrouting_obj = self.env['finishfetplanmodule.jobroutingtable']
                jobrouting_id = jobrouting_obj.search([('colour', '=', str(getcol.fill)[139:147])])
                for thisjob in jobrouting_id:
                    # add record  for Shift B
                    thisheader_ids.actualitemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id, 'jobrouting_id': thisjob.id,
                         'date': getdt,
                         'name': 'Added Shift B',
                         'shift_a': 0,
                         'shift_b': getcol.value,
                         'shift_c': 0})

                # Reading for Shift C
                getdt = self.from_dt + timedelta(relativedate)
                getcol = worksheet.cell(row=itempos, column=i + 2)
                jobrouting_obj = self.env['finishfetplanmodule.jobroutingtable']
                jobrouting_id = jobrouting_obj.search([('colour', '=', str(getcol.fill)[139:147])])

                for thisjob in jobrouting_id:
                    # add record  for Shift C
                    thisheader_ids.actualitemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id, 'jobrouting_id': thisjob.id,
                         'date': getdt,
                         'name': 'Added Shift C',
                         'shift_a': 0,
                         'shift_b': 0,
                         'shift_c': getcol.value})

                relativedate = relativedate + 1
            itempos = itempos + 2
        self.readfromexcel = 'SAVE data for future reference'
