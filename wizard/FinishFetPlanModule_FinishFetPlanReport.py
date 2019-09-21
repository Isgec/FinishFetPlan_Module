from odoo import models, fields
import io
import base64
import openpyxl
from datetime import timedelta
from os import path
from openpyxl.styles import PatternFill
from tempfile import TemporaryFile
from openpyxl.comments import Comment


class FinishFetPlanModule_FinishFetPlanReport(models.TransientModel):
    _name = 'finishfetplanmodule.finishfetplanreport'
    from_dt = fields.Date('Step 1: Set the Date', required=True)
    name = fields.Char('Finish Fet Excel Sheet')
    upload_file = fields.Binary(string="Step 2: Upload Excel File having Plan and Actual")
    uploadedfilename = fields.Char('File Name', size=256, default='Select Revised Plan')
    readfromexcel = fields.Text('Final Step')
    download_file = fields.Binary(string="Step 4: Download Generated Plan and Review the Load")
    downloadedfilename = fields.Char('File Name', size=256, default='Download Generated Plan.xlsx')
    report_flag = fields.Integer('Report Genrated Flag')
    remarks = fields.Text('Remarks')

    def upload_excel(self, data, context=None):
        # Generating of the excel file to be read by openpyxl
        file = base64.decodestring(self.upload_file)
        excel_fileobj = TemporaryFile('wb+')
        excel_fileobj.write(file)
        excel_fileobj.seek(0)

        # Create workbook
        wb = openpyxl.load_workbook(excel_fileobj, data_only=True)
        wb.active = 0
        worksheet = wb.active

        header_obj = self.env['finishfetplanmodule.itemplanheadertable']
        header_ids = header_obj.search([(1, '=', 1)])
        self.readfromexcel = 'Start'
        itempos = 26
        relativedate = 0
        my_max_col = 120
        for thisheader_ids in header_ids:
            self.readfromexcel = self.readfromexcel + '{ Items : ' + thisheader_ids.name + '}'
            relativedate = 0
            my_max_col = 120
            for thisitems_ids in thisheader_ids.itemplan_id:
                if thisitems_ids.date >= self.from_dt:
                    jobroutingid = thisitems_ids.jobrouting_id.id
                    plandate = thisitems_ids.date
                    # self.readfromexcel = self.readfromexcel + ', Unlinked->' + str(thisitems_ids.date)
                    thisitems_ids.unlink()  # Delete Record  from Item table With particular

            for i in range(4, my_max_col + 1, 3):
                # Reading for Shift A
                getdt = self.from_dt + timedelta(relativedate)
                getcol = worksheet.cell(row=itempos, column=i)
                jobrouting_obj = self.env['finishfetplanmodule.jobroutingtable']
                jobrouting_id = jobrouting_obj.search([('colour', '=', str(getcol.fill)[139:147])])
                self.readfromexcel = self.readfromexcel + ' {Reading Cell Shift A:' + str(getcol.fill) + '} '
                setrouting_id = None
                for thisjob in jobrouting_id:
                    setrouting_id = thisjob.id
                if setrouting_id:
                    thisheader_ids.itemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id, 'jobrouting_id': setrouting_id,
                         'date': getdt,
                         'name': 'Added Shift A',
                         'shift_a_c': getcol.value,
                         'shift_b_c': '',
                         'shift_c_c': ''})
                else:
                    thisheader_ids.itemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id,
                         'date': getdt,
                         'name': 'Added TextValue in A',
                         'shift_a_c': getcol.value,
                         'shift_b_c': '',
                         'shift_c_c': ''})
                # Reading for Shift B
                getdt = self.from_dt + timedelta(relativedate)
                getcol = worksheet.cell(row=itempos, column=i + 1)
                jobrouting_obj = self.env['finishfetplanmodule.jobroutingtable']
                jobrouting_id = jobrouting_obj.search([('colour', '=', str(getcol.fill)[139:147])])
                # self.readfromexcel = self.readfromexcel + ' {Reading Cell Shift B:' + str(getcol.fill)[139:147]+ '} '
                setrouting_id = None
                for thisjob in jobrouting_id:
                    setrouting_id = thisjob.id
                if setrouting_id:
                    thisheader_ids.itemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id, 'jobrouting_id': setrouting_id,
                         'date': getdt,
                         'name': 'Added Shift B',
                         'shift_a_c': '',
                         'shift_b_c': getcol.value,
                         'shift_c_c': ''})
                else:
                    thisheader_ids.itemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id,
                         'date': getdt,
                         'name': 'Added TextValue in B',
                         'shift_a_c': '',
                         'shift_b_c': getcol.value,
                         'shift_c_c': ''})

                # Reading for Shift C
                getdt = self.from_dt + timedelta(relativedate)
                getcol = worksheet.cell(row=itempos, column=i + 2)
                jobrouting_obj = self.env['finishfetplanmodule.jobroutingtable']
                jobrouting_id = jobrouting_obj.search([('colour', '=', str(getcol.fill)[139:147])])
                # self.readfromexcel = self.readfromexcel + ' {Reading Cell Shift C:' + str(getcol.fill)[139:147]+ '} '
                setrouting_id = None
                for thisjob in jobrouting_id:
                    setrouting_id = thisjob.id
                if setrouting_id:
                    thisheader_ids.itemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id, 'jobrouting_id': setrouting_id,
                         'date': getdt,
                         'name': 'Added Shift C',
                         'shift_a_c': '',
                         'shift_b_c': '',
                         'shift_c_c': getcol.value})
                else:
                    thisheader_ids.itemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id,
                         'date': getdt,
                         'name': 'Added TextValue in C',
                         'shift_a_c': '',
                         'shift_b_c': '',
                         'shift_c_c': getcol.value})

                relativedate = relativedate + 1
            itempos = itempos + 2

        #  Update Actual Item
        header_obj = self.env['finishfetplanmodule.itemplanheadertable']
        header_ids = header_obj.search([(1, '=', 1)])
        # self.readfromexcel = ''
        itempos = 27
        relativedate = 0
        my_max_col = 120
        for thisheader_ids in header_ids:
            # self.readfromexcel = self.readfromexcel + '{ Items : ' + thisheader_ids.name + '}'
            relativedate = 0
            my_max_col = 120
            for thisitems_ids in thisheader_ids.actualitemplan_id:
                if thisitems_ids.date >= self.from_dt:
                    jobroutingid = thisitems_ids.jobrouting_id.id
                    plandate = thisitems_ids.date
                    thisitems_ids.unlink()  # Delete Record  from Item table With particular

            for i in range(4, my_max_col + 1, 3):
                # Reading for Shift A
                getdt = self.from_dt + timedelta(relativedate)
                getcol = worksheet.cell(row=itempos, column=i)
                jobrouting_obj = self.env['finishfetplanmodule.jobroutingtable']
                jobrouting_id = jobrouting_obj.search([('colour', '=', str(getcol.fill)[139:147])])
                setrouting_id = None
                for thisjob in jobrouting_id:
                    setrouting_id = thisjob.id
                if setrouting_id:
                    thisheader_ids.actualitemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id, 'jobrouting_id': setrouting_id,
                         'date': getdt,
                         'name': 'Added Text in Shift A',
                         'shift_a_c': getcol.value,
                         'shift_b_c': '',
                         'shift_c_c': ''})
                else:
                    thisheader_ids.actualitemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id,
                         'date': getdt,
                         'name': 'Added TextValue in A',
                         'shift_a_c': getcol.value,
                         'shift_b_c': '',
                         'shift_c_c': ''})

                # Reading for Shift B
                getdt = self.from_dt + timedelta(relativedate)
                getcol = worksheet.cell(row=itempos, column=i + 1)
                jobrouting_obj = self.env['finishfetplanmodule.jobroutingtable']
                jobrouting_id = jobrouting_obj.search([('colour', '=', str(getcol.fill)[139:147])])
                setrouting_id = None
                for thisjob in jobrouting_id:
                    setrouting_id = thisjob.id
                if setrouting_id:
                    # add record  for Shift B
                    thisheader_ids.actualitemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id, 'jobrouting_id': setrouting_id,
                         'date': getdt,
                         'name': 'Added Shift B',
                         'shift_a_c': '',
                         'shift_b_c': getcol.value,
                         'shift_c_c': ''})
                else:
                    thisheader_ids.actualitemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id,
                         'date': getdt,
                         'name': 'Added TextValue in B',
                         'shift_a_c': '',
                         'shift_b_c': getcol.value,
                         'shift_c_c': ''})

                # Reading for Shift C
                getdt = self.from_dt + timedelta(relativedate)
                getcol = worksheet.cell(row=itempos, column=i + 2)
                jobrouting_obj = self.env['finishfetplanmodule.jobroutingtable']
                jobrouting_id = jobrouting_obj.search([('colour', '=', str(getcol.fill)[139:147])])
                setrouting_id = None
                for thisjob in jobrouting_id:
                    setrouting_id = thisjob.id
                if setrouting_id:
                    # add record  for Shift C
                    thisheader_ids.actualitemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id, 'jobrouting_id': setrouting_id,
                         'date': getdt,
                         'name': 'Added Shift C',
                         'shift_a_c': '',
                         'shift_b_c': '',
                         'shift_c_c': getcol.value})
                else:
                    thisheader_ids.actualitemplan_id.create(
                        {'itemplanheader_id': thisheader_ids.id,
                         'date': getdt,
                         'name': 'Added C text value',
                         'shift_a_c': '',
                         'shift_b_c': '',
                         'shift_c_c': getcol.value})


                relativedate = relativedate + 1
            itempos = itempos + 2
        # self.button_excel(data, context=None)
        self.readfromexcel = 'Step 5: You may review the Impact on Load and Re-plan or SAVE and data for future reference'

    def button_excel(self, data, context=None):
        fillGRINDING = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        fillGOUGING = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        fillWELDING = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        commGRINDING = [None] * 200
        commGOUGING = [None] * 200
        commWELDING = [None] * 200
        src = path.dirname(path.realpath(__file__)) + "/FinishFetplan.xlsx"
        wb = openpyxl.load_workbook(src)
        wb.active = 0
        worksheet = wb.active
        filename = 'FinishFetplan.xlsx'
        reportname = "FinishFetPlan:"
        self.name = 'Finish Fet Plan Report as on: ' + str(self.from_dt)

        # Start of Section to generate load portion of sheet
        wb.active = 0
        worksheet = wb.active
        itemheader_obj = self.env['finishfetplanmodule.manpowertable']
        item_ids = itemheader_obj.search([(1, '=', 1)])
        for thisitem in item_ids:
            if thisitem.jobrouting_id.name == 'WELDING':
                welding_shift_a = thisitem.shift_a
                welding_shift_b = thisitem.shift_b
                welding_shift_c = thisitem.shift_c
            if thisitem.jobrouting_id.name == 'GRINDING':
                grinding_shift_a = thisitem.shift_a
                grinding_shift_b = thisitem.shift_b
                grinding_shift_c = thisitem.shift_c
            if thisitem.jobrouting_id.name == 'Gouging':
                gouging_shift_a = thisitem.shift_a
                gouging_shift_b = thisitem.shift_b
                gouging_shift_c = thisitem.shift_c

        itemheader_obj = self.env['finishfetplanmodule.itemplantable']
        item_ids = itemheader_obj.search([('date', '>=', self.from_dt)])

        for thisitem in item_ids:
            if thisitem.date >= self.from_dt:
                datediff = (thisitem.date - self.from_dt)
                col = ((datediff.days + 1) * 3) + 1
                dateStr = str(thisitem.date.day) + "/" + str(thisitem.date.month)
                setdate = worksheet.cell(row=11, column=col)
                setdate.value = dateStr
                if thisitem.jobrouting_id.name == 'WELDING':
                    row = 14
                    if thisitem.shift_a > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = welding_shift_a
                        setcol2 = worksheet.cell(row=row, column=col)
                        if setcol2.value:
                            setcol2.value = setcol2.value + thisitem.shift_a or ''
                            # Setting value of comment if WELDER is Overloaded in Shift A
                            if welding_shift_a < setcol2.value:
                                commWELDING[col] = Comment(
                                    'Overloaded: Out of ' + str(welding_shift_a) + ' WELDER in Shift A loaded ' + str(
                                        setcol2.value), 'System')
                            # END of : Setting value of comment if WELDER is Overloaded in Shift A
                        else:
                            setcol2.value = thisitem.shift_a or ''
                    col = col + 1

                    if thisitem.shift_b > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = welding_shift_b
                        setcol3 = worksheet.cell(row=row, column=col)
                        if setcol3.value:
                            setcol3.value = setcol3.value + thisitem.shift_b or ''
                            # Setting value of comment if WELDER is Overloaded in Shift B
                            if welding_shift_b < setcol3.value:
                                commWELDING[col] = Comment(
                                    'Overloaded: Out of ' + str(welding_shift_b) + ' WELDER in Shift B loaded ' + str(
                                        setcol3.value), 'System')
                            # END of : Setting value of comment if WELDER is Overloaded in Shift B

                        else:
                            setcol3.value = thisitem.shift_b or ''
                    col = col + 1

                    if thisitem.shift_c > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = welding_shift_c
                        setcol4 = worksheet.cell(row=row, column=col)
                        if setcol4.value:
                            setcol4.value = setcol4.value + thisitem.shift_c
                            # Setting value of comment if WELDER is Overloaded in Shift C
                            if welding_shift_c < setcol4.value:
                                commWELDING[col] = Comment(
                                    'Overloaded: Out of ' + str(welding_shift_c) + ' WELDER in Shift C loaded ' + str(
                                        setcol4.value), 'System')
                            # END of : Setting value of comment if WELDER is Overloaded in Shift A

                        else:
                            setcol4.value = thisitem.shift_c

                if thisitem.jobrouting_id.name == 'GRINDING':
                    row = 17
                    if thisitem.shift_a > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = grinding_shift_a
                        setcol2 = worksheet.cell(row=row, column=col)
                        if setcol2.value:
                            setcol2.value = setcol2.value + thisitem.shift_a or ''
                            # Setting value of comment if GRINDING is Overloaded in Shift A
                            if grinding_shift_a < setcol2.value:
                                commGRINDING[col] = Comment(
                                    'Overloaded: Out of ' + str(
                                        grinding_shift_a) + ' GRINDING in Shift A loaded ' + str(
                                        setcol2.value), 'System')
                            # END of : Setting value of comment if WELDER is Overloaded in Shift A
                        else:
                            setcol2.value = thisitem.shift_a or ''
                    col = col + 1

                    if thisitem.shift_b > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = grinding_shift_b
                        setcol3 = worksheet.cell(row=row, column=col)
                        if setcol3.value:
                            setcol3.value = setcol3.value + thisitem.shift_b or ''
                            # Setting value of comment if GRINDING is Overloaded in Shift B
                            if grinding_shift_b < setcol3.value:
                                commGRINDING[col] = Comment(
                                    'Overloaded: Out of ' + str(
                                        grinding_shift_b) + ' GRINDING in Shift B loaded ' + str(
                                        setcol3.value), 'System')
                            # END of : Setting value of comment if GRINDING is Overloaded in Shift B
                        else:
                            setcol3.value = thisitem.shift_b or ''
                    col = col + 1

                    if thisitem.shift_c > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = grinding_shift_c
                        setcol4 = worksheet.cell(row=row, column=col)
                        if setcol4.value:
                            setcol4.value = setcol4.value + thisitem.shift_c
                            # Setting value of comment if GRINDING is Overloaded in Shift C
                            if grinding_shift_c < setcol4.value:
                                commGRINDING[col] = Comment(
                                    'Overloaded: Out of ' + str(
                                        grinding_shift_c) + ' GRINDING in Shift C loaded ' + str(
                                        setcol4.value), 'System')
                            # END of : Setting value of comment if GRINDING is Overloaded in Shift C

                        else:
                            setcol4.value = thisitem.shift_c

                if thisitem.jobrouting_id.name == 'Gouging':
                    row = 20
                    if thisitem.shift_a > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = gouging_shift_a
                        setcol2 = worksheet.cell(row=row, column=col)
                        if setcol2.value:
                            setcol2.value = setcol2.value + thisitem.shift_a or ''
                            # Setting value of comment if Gouging is Overloaded in Shift A
                            if gouging_shift_a < setcol2.value:
                                commGOUGING[col] = Comment(
                                    'Overloaded: Out of ' + str(
                                        gouging_shift_a) + ' Gouging in Shift A loaded ' + str(
                                        setcol2.value), 'System')
                            # END of : Setting value of comment if Gouging is Overloaded in Shift A
                        else:
                            setcol2.value = thisitem.shift_a or ''
                    col = col + 1

                    if thisitem.shift_b > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = gouging_shift_b
                        setcol3 = worksheet.cell(row=row, column=col)
                        if setcol3.value:
                            setcol3.value = setcol3.value + thisitem.shift_b or ''
                            # Setting value of comment if Gouging is Overloaded in Shift B
                            if gouging_shift_b < setcol3.value:
                                commGOUGING[col] = Comment(
                                    'Overloaded: Out of ' + str(
                                        gouging_shift_b) + ' Gouging in Shift B loaded ' + str(
                                        setcol3.value), 'System')
                            # END of : Setting value of comment if Gouging is Overloaded in Shift B
                        else:
                            setcol3.value = thisitem.shift_b or ''
                    col = col + 1

                    if thisitem.shift_c > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = gouging_shift_c
                        setcol4 = worksheet.cell(row=row, column=col)
                        if setcol4.value:
                            setcol4.value = setcol4.value + thisitem.shift_c
                            # Setting value of comment if Gouging is Overloaded in Shift C
                            if gouging_shift_c < setcol4.value:
                                commGOUGING[col] = Comment(
                                    'Overloaded: Out of ' + str(
                                        gouging_shift_c) + ' Gouging in Shift C loaded ' + str(
                                        setcol4.value), 'System')
                            # END of : Setting value of comment if Gouging is Overloaded in Shift C
                        else:
                            setcol4.value = thisitem.shift_c
        ## End of Generation of Load portion of sheet
        # Start of Section to Generate Plan portion of the sheet
        itemheader_obj = self.env['finishfetplanmodule.itemplanheadertable']
        itemheader_ids = itemheader_obj.search([(1, '=', 1)])
        row = 26
        for thisitems_id in itemheader_ids:
            setcol1 = worksheet.cell(row=row, column=1)
            setcol1.value = thisitems_id.name or ''
            setcol1 = worksheet.cell(row=row, column=2)
            setcol1.value = thisitems_id.wo_srno or ''

            for thisitem in thisitems_id.itemplan_id:
                if thisitem.date >= self.from_dt:
                    datediff = (thisitem.date - self.from_dt)
                    col = ((datediff.days + 1) * 3) + 1
                    dateStr = str(thisitem.date.day) + "/" + str(thisitem.date.month)
                    setdate = worksheet.cell(row=24, column=col)
                    setdate.value = dateStr

                    # for  String  Value Shift A
                    if thisitem.shift_a_c != '':
                        setcol2 = worksheet.cell(row=row, column=col)
                        setcol2.value = thisitem.shift_a_c or ''

                    # for  Decimal or Number  Value Shift A
                    if thisitem.shift_a > 0:
                        setcol2 = worksheet.cell(row=row, column=col)
                        setcol2.value = thisitem.shift_a or ''
                        colorfill = PatternFill(start_color=thisitem.jobrouting_id.colour,
                                                end_color=thisitem.jobrouting_id.colour, fill_type='solid')

                        setcol2.fill = colorfill
                        # Setting value of comment if WELDER is Overloaded in Shift A
                        if thisitem.jobrouting_id.name == 'WELDING':
                            if commWELDING[col]:
                                setcol2.comment = commWELDING[col]
                        # END of Setting value of comment if WELDER is Overloaded in Shift A

                        # Setting value of comment if GRINDING is Overloaded in Shift A
                        if thisitem.jobrouting_id.name == 'GRINDING':
                            if commGRINDING[col]:
                                setcol2.comment = commGRINDING[col]
                        # END of Setting value of comment if GRINDING is Overloaded in Shift A

                        # Setting value of comment if Gouging is Overloaded in Shift A
                        if thisitem.jobrouting_id.name == 'Gouging':
                            if commGOUGING[col]:
                                setcol2.comment = commGOUGING[col]
                        # END of Setting value of comment if Gouging is Overloaded in Shift A


                    col = col + 1

                    # for  String  Value Shift B
                    if thisitem.shift_b_c != '':
                        setcol3 = worksheet.cell(row=row, column=col)
                        setcol3.value = thisitem.shift_b_c or ''

                    # for  Decimal or Number  Value Shift B
                    if thisitem.shift_b > 0:
                        setcol3 = worksheet.cell(row=row, column=col)
                        setcol3.value = thisitem.shift_b or ''
                        colorfill = PatternFill(start_color=thisitem.jobrouting_id.colour,
                                                end_color=thisitem.jobrouting_id.colour, fill_type='solid')

                        setcol3.fill = colorfill
                        # Setting value of comment if WELDER is Overloaded in Shift B
                        if thisitem.jobrouting_id.name == 'WELDING':
                            if commWELDING[col]:
                                setcol3.comment = commWELDING[col]
                        # END of Setting value of comment if WELDER is Overloaded in Shift B

                        # Setting value of comment if GRINDING is Overloaded in Shift B
                        if thisitem.jobrouting_id.name == 'GRINDING':
                            if commGRINDING[col]:
                                setcol3.comment = commGRINDING[col]
                        # END of Setting value of comment if GRINDING is Overloaded in Shift B

                        # Setting value of comment if Gouging is Overloaded in Shift B
                        if thisitem.jobrouting_id.name == 'Gouging':
                            if commGOUGING[col]:
                                setcol3.comment = commGOUGING[col]
                        # END of Setting value of comment if Gouging is Overloaded in Shift B

                    col = col + 1

                    # for  String  Value Shift C
                    if thisitem.shift_c_c != '':
                        setcol4 = worksheet.cell(row=row, column=col)
                        setcol4.value = thisitem.shift_c_c or ''

                    # for  Decimal or Number  Value Shift C
                    if thisitem.shift_c > 0:
                        setcol4 = worksheet.cell(row=row, column=col)
                        setcol4.value = thisitem.shift_c
                        colorfill = PatternFill(start_color=thisitem.jobrouting_id.colour,
                                                end_color=thisitem.jobrouting_id.colour, fill_type='solid')

                        setcol4.fill = colorfill
                        # Setting value of comment if WELDER is Overloaded in Shift C
                        if thisitem.jobrouting_id.name == 'WELDING':
                            if commWELDING[col]:
                                setcol4.comment = commWELDING[col]
                        # END of Setting value of comment if WELDER is Overloaded in Shift C

                        # Setting value of comment if GRINDING is Overloaded in Shift B
                        if thisitem.jobrouting_id.name == 'GRINDING':
                            if commGRINDING[col]:
                                setcol4.comment = commGRINDING[col]
                        # END of Setting value of comment if GRINDING is Overloaded in Shift C

                        # Setting value of comment if Gouging is Overloaded in Shift C
                        if thisitem.jobrouting_id.name == 'Gouging':
                            if commGOUGING[col]:
                                setcol4.comment = commGOUGING[col]
                        # END of Setting value of comment if Gouging is Overloaded in Shift C

            row = row + 2
        ## End of Section to generate Plan portion of the sheet

        # Start of Generation of Actual portion of sheet
        wb.active = 0
        worksheet = wb.active
        # Generating Actual Portion

        itemheader_obj = self.env['finishfetplanmodule.itemplanheadertable']
        itemheader_ids = itemheader_obj.search([(1, '=', 1)])
        row = 27
        for thisitems_id in itemheader_ids:
            # Thse are already set while writing Plan portion. Row 15 is merged with Row 14 so write not possible
            # setcol1 = worksheet.cell(row=row, column=1)
            # setcol1.value = thisitems_id.name or ''
            # setcol1 = worksheet.cell(row=row, column=2)
            # setcol1.value = thisitems_id.wo_srno or ''

            for thisitem in thisitems_id.actualitemplan_id:
                if thisitem.date >= self.from_dt:
                    datediff = (thisitem.date - self.from_dt)
                    col = ((datediff.days + 1) * 3) + 1
                    dateStr = str(thisitem.date.day) + "/" + str(thisitem.date.month)
                    setdate = worksheet.cell(row=24, column=col)
                    setdate.value = dateStr

                    # for  String  Value Shift A
                    if thisitem.shift_a_c != '':
                        setcol2 = worksheet.cell(row=row, column=col)
                        setcol2.value = thisitem.shift_a_c or ''
                    # for  Decimal or Number  Value Shift A

                    if thisitem.shift_a > 0:
                        setcol2 = worksheet.cell(row=row, column=col)
                        setcol2.value = thisitem.shift_a or ''
                        colorfill = PatternFill(start_color=thisitem.jobrouting_id.colour,
                                                end_color=thisitem.jobrouting_id.colour, fill_type='solid')

                        setcol2.fill = colorfill
                    col = col + 1
                    # for  String  Value Shift B
                    if thisitem.shift_b_c != '':
                        setcol3 = worksheet.cell(row=row, column=col)
                        setcol3.value = thisitem.shift_b_c or ''

                    # for  Decimal or Number  Value Shift B
                    if thisitem.shift_b > 0:
                        setcol3 = worksheet.cell(row=row, column=col)
                        setcol3.value = thisitem.shift_b or ''
                        colorfill = PatternFill(start_color=thisitem.jobrouting_id.colour,
                                                end_color=thisitem.jobrouting_id.colour, fill_type='solid')

                        setcol3.fill = colorfill
                    col = col + 1
                    # for  String  Value Shift C
                    if thisitem.shift_c_c != '':
                        setcol4 = worksheet.cell(row=row, column=col)
                        setcol4.value = thisitem.shift_c_c or ''

                    # for  Decimal or Number  Value Shift C
                    if thisitem.shift_c > 0:
                        setcol4 = worksheet.cell(row=row, column=col)
                        setcol4.value = thisitem.shift_c
                        colorfill = PatternFill(start_color=thisitem.jobrouting_id.colour,
                                                end_color=thisitem.jobrouting_id.colour, fill_type='solid')

                        setcol4.fill = colorfill
            row = row + 2
        wb.active = 0
        worksheet = wb.active
        fp = io.BytesIO()
        wb.save(fp)
        out = base64.encodestring(fp.getvalue())
        self.download_file = out
        self.report_flag = 1
