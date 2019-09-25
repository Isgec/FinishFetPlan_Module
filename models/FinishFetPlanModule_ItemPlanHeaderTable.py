from odoo import models, fields, api
import io
import base64
import openpyxl
from datetime import timedelta
from os import path
from openpyxl.styles import PatternFill
from tempfile import TemporaryFile
from openpyxl.comments import Comment


class FinishFetPlanModuleItemPlanHeaderTable(models.Model):
    _name = 'finishfetplanmodule.itemplanheadertable'

    itemplan_id = fields.One2many('finishfetplanmodule.itemplantable', 'itemplanheader_id')
    actualitemplan_id = fields.One2many('finishfetplanmodule.actualitemplantable', 'itemplanheader_id')
    name = fields.Char('Item ', required=True)
    wo_srno = fields.Char('W.O/ SNo. ', required=True)
    plan_date = fields.Date('Plan Date ', required=True)
    item_status = fields.Boolean(string="Deactivated", default=True)

    def rescheduledate(self, data, context=None):
        for record in self.itemplan_id:
            if record:
                record.date = self.plan_date + timedelta(days=record.lag_days)

    def rescheduledate2(self, data, context=None):
        for record in self.actualitemplan_id:
            if record:
                record.date = self.plan_date + timedelta(days=record.lag_days)

    def button_excel(self, data, context=None):
        fillGRINDING = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        fillGOUGING = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        fillWELDING = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        commGRINDING = [None] * 200
        commGOUGING = [None] * 200
        commWELDING = [None] * 200
        src = path.realpath("/home/rajeev/myaddons/FinishFetPlan_Module/wizard/FinishFetplan.xlsx")
        wb = openpyxl.load_workbook(src)
        wb.active = 0
        worksheet = wb.active
        filename = self.name + '.xlsx'
        reportname = "Items : " + self.name
        # self.name = 'Finish Fet Plan Report as on: ' + str(self.plan_date)

        # Start of Section to generate load portion of sheet
        wb.active = 0
        worksheet = wb.active
        itemheader_obj = self.env['finishfetplanmodule.manpowertable']
        item_ids = itemheader_obj.search([(1, '=', 1)])
        for thisitem in item_ids:
            if thisitem.jobrouting_id.name == 'WELDING':
                welding_shift_a = thisitem.shift_a
                welding_shift_b = thisitem.shift_b
                welding_shift_c = thisitem.shift_c
            if thisitem.jobrouting_id.name == 'GRINDING':
                grinding_shift_a = thisitem.shift_a
                grinding_shift_b = thisitem.shift_b
                grinding_shift_c = thisitem.shift_c
            if thisitem.jobrouting_id.name == 'Gouging':
                gouging_shift_a = thisitem.shift_a
                gouging_shift_b = thisitem.shift_b
                gouging_shift_c = thisitem.shift_c

        itemheader_obj = self.env['finishfetplanmodule.itemplantable']
        item_ids = itemheader_obj.search([('date', '>=', self.plan_date)])

        for thisitem in item_ids:
            if thisitem.date >= self.plan_date:
                datediff = (thisitem.date - self.plan_date)
                col = ((datediff.days + 1) * 3) + 1
                dateStr = str(thisitem.date.day) + "/" + str(thisitem.date.month)
                setdate = worksheet.cell(row=11, column=col)
                setdate.value = dateStr
                if thisitem.jobrouting_id.name == 'WELDING':
                    row = 14
                    if thisitem.shift_a > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = welding_shift_a
                        setcol2 = worksheet.cell(row=row, column=col)
                        if setcol2.value:
                            setcol2.value = setcol2.value + thisitem.shift_a or ''
                            # Setting value of comment if WELDER is Overloaded in Shift A
                            if welding_shift_a < setcol2.value:
                                commWELDING[col] = Comment(
                                    'Overloaded: Out of ' + str(welding_shift_a) + ' WELDER in Shift A loaded ' + str(
                                        setcol2.value), 'System')
                            # END of : Setting value of comment if WELDER is Overloaded in Shift A
                        else:
                            setcol2.value = thisitem.shift_a or ''
                    col = col + 1

                    if thisitem.shift_b > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = welding_shift_b
                        setcol3 = worksheet.cell(row=row, column=col)
                        if setcol3.value:
                            setcol3.value = setcol3.value + thisitem.shift_b or ''
                            # Setting value of comment if WELDER is Overloaded in Shift B
                            if welding_shift_b < setcol3.value:
                                commWELDING[col] = Comment(
                                    'Overloaded: Out of ' + str(welding_shift_b) + ' WELDER in Shift B loaded ' + str(
                                        setcol3.value), 'System')
                            # END of : Setting value of comment if WELDER is Overloaded in Shift B

                        else:
                            setcol3.value = thisitem.shift_b or ''
                    col = col + 1

                    if thisitem.shift_c > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = welding_shift_c
                        setcol4 = worksheet.cell(row=row, column=col)
                        if setcol4.value:
                            setcol4.value = setcol4.value + thisitem.shift_c
                            # Setting value of comment if WELDER is Overloaded in Shift C
                            if welding_shift_c < setcol4.value:
                                commWELDING[col] = Comment(
                                    'Overloaded: Out of ' + str(welding_shift_c) + ' WELDER in Shift C loaded ' + str(
                                        setcol4.value), 'System')
                            # END of : Setting value of comment if WELDER is Overloaded in Shift A

                        else:
                            setcol4.value = thisitem.shift_c

                if thisitem.jobrouting_id.name == 'GRINDING':
                    row = 17
                    if thisitem.shift_a > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = grinding_shift_a
                        setcol2 = worksheet.cell(row=row, column=col)
                        if setcol2.value:
                            setcol2.value = setcol2.value + thisitem.shift_a or ''
                            # Setting value of comment if GRINDING is Overloaded in Shift A
                            if grinding_shift_a < setcol2.value:
                                commGRINDING[col] = Comment(
                                    'Overloaded: Out of ' + str(
                                        grinding_shift_a) + ' GRINDING in Shift A loaded ' + str(
                                        setcol2.value), 'System')
                            # END of : Setting value of comment if WELDER is Overloaded in Shift A
                        else:
                            setcol2.value = thisitem.shift_a or ''
                    col = col + 1

                    if thisitem.shift_b > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = grinding_shift_b
                        setcol3 = worksheet.cell(row=row, column=col)
                        if setcol3.value:
                            setcol3.value = setcol3.value + thisitem.shift_b or ''
                            # Setting value of comment if GRINDING is Overloaded in Shift B
                            if grinding_shift_b < setcol3.value:
                                commGRINDING[col] = Comment(
                                    'Overloaded: Out of ' + str(
                                        grinding_shift_b) + ' GRINDING in Shift B loaded ' + str(
                                        setcol3.value), 'System')
                            # END of : Setting value of comment if GRINDING is Overloaded in Shift B
                        else:
                            setcol3.value = thisitem.shift_b or ''
                    col = col + 1

                    if thisitem.shift_c > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = grinding_shift_c
                        setcol4 = worksheet.cell(row=row, column=col)
                        if setcol4.value:
                            setcol4.value = setcol4.value + thisitem.shift_c
                            # Setting value of comment if GRINDING is Overloaded in Shift C
                            if grinding_shift_c < setcol4.value:
                                commGRINDING[col] = Comment(
                                    'Overloaded: Out of ' + str(
                                        grinding_shift_c) + ' GRINDING in Shift C loaded ' + str(
                                        setcol4.value), 'System')
                            # END of : Setting value of comment if GRINDING is Overloaded in Shift C

                        else:
                            setcol4.value = thisitem.shift_c

                if thisitem.jobrouting_id.name == 'Gouging':
                    row = 20
                    if thisitem.shift_a > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = gouging_shift_a
                        setcol2 = worksheet.cell(row=row, column=col)
                        if setcol2.value:
                            setcol2.value = setcol2.value + thisitem.shift_a or ''
                            # Setting value of comment if Gouging is Overloaded in Shift A
                            if gouging_shift_a < setcol2.value:
                                commGOUGING[col] = Comment(
                                    'Overloaded: Out of ' + str(
                                        gouging_shift_a) + ' Gouging in Shift A loaded ' + str(
                                        setcol2.value), 'System')
                            # END of : Setting value of comment if Gouging is Overloaded in Shift A
                        else:
                            setcol2.value = thisitem.shift_a or ''
                    col = col + 1

                    if thisitem.shift_b > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = gouging_shift_b
                        setcol3 = worksheet.cell(row=row, column=col)
                        if setcol3.value:
                            setcol3.value = setcol3.value + thisitem.shift_b or ''
                            # Setting value of comment if Gouging is Overloaded in Shift B
                            if gouging_shift_b < setcol3.value:
                                commGOUGING[col] = Comment(
                                    'Overloaded: Out of ' + str(
                                        gouging_shift_b) + ' Gouging in Shift B loaded ' + str(
                                        setcol3.value), 'System')
                            # END of : Setting value of comment if Gouging is Overloaded in Shift B
                        else:
                            setcol3.value = thisitem.shift_b or ''
                    col = col + 1

                    if thisitem.shift_c > 0:
                        setcol = worksheet.cell(row=row - 1, column=col)
                        setcol.value = gouging_shift_c
                        setcol4 = worksheet.cell(row=row, column=col)
                        if setcol4.value:
                            setcol4.value = setcol4.value + thisitem.shift_c
                            # Setting value of comment if Gouging is Overloaded in Shift C
                            if gouging_shift_c < setcol4.value:
                                commGOUGING[col] = Comment(
                                    'Overloaded: Out of ' + str(
                                        gouging_shift_c) + ' Gouging in Shift C loaded ' + str(
                                        setcol4.value), 'System')
                            # END of : Setting value of comment if Gouging is Overloaded in Shift C
                        else:
                            setcol4.value = thisitem.shift_c

        ## End of Generation of Load portion of sheet
        # Start of Section to Generate Plan portion of the sheet
        itemheader_obj = self.env['finishfetplanmodule.itemplanheadertable']
        itemheader_ids = itemheader_obj.search(
            ['&', ('wo_srno', '=', self.wo_srno), ('plan_date', '=', self.plan_date)])
        row = 26
        for thisitems_id in itemheader_ids:
            setcol1 = worksheet.cell(row=row, column=1)
            setcol1.value = thisitems_id.name or ''
            setcol1 = worksheet.cell(row=row, column=2)
            setcol1.value = thisitems_id.wo_srno or ''

            for thisitem in thisitems_id.itemplan_id:
                if thisitem.date >= self.plan_date:
                    datediff = (thisitem.date - self.plan_date)
                    col = ((datediff.days + 1) * 3) + 1
                    dateStr = str(thisitem.date.day) + "/" + str(thisitem.date.month)
                    setdate = worksheet.cell(row=24, column=col)
                    setdate.value = dateStr

                    # for  String  Value Shift A
                    if thisitem.shift_a_c != '':
                        setcol2 = worksheet.cell(row=row, column=col)
                        setcol2.value = thisitem.shift_a_c or ''
                        colorfill = PatternFill(start_color=thisitem.bg_color_cell,
                                                end_color=thisitem.bg_color_cell, fill_type='solid')
                        setcol2.fill = colorfill

                    # for  Decimal or Number  Value Shift A
                    if thisitem.shift_a > 0 and thisitem.jobrouting_id:
                        setcol2 = worksheet.cell(row=row, column=col)
                        setcol2.value = thisitem.shift_a or ''
                        colorfill = PatternFill(start_color=thisitem.jobrouting_id.colour,
                                                end_color=thisitem.jobrouting_id.colour, fill_type='solid')

                        setcol2.fill = colorfill
                        # Setting value of comment if WELDER is Overloaded in Shift A
                        if thisitem.jobrouting_id.name == 'WELDING':
                            if commWELDING[col]:
                                setcol2.comment = commWELDING[col]
                        # END of Setting value of comment if WELDER is Overloaded in Shift A

                        # Setting value of comment if GRINDING is Overloaded in Shift A
                        if thisitem.jobrouting_id.name == 'GRINDING':
                            if commGRINDING[col]:
                                setcol2.comment = commGRINDING[col]
                        # END of Setting value of comment if GRINDING is Overloaded in Shift A

                        # Setting value of comment if Gouging is Overloaded in Shift A
                        if thisitem.jobrouting_id.name == 'Gouging':
                            if commGOUGING[col]:
                                setcol2.comment = commGOUGING[col]
                        # END of Setting value of comment if Gouging is Overloaded in Shift A

                    col = col + 1

                    # for  String  Value Shift B
                    if thisitem.shift_b_c != '':
                        setcol3 = worksheet.cell(row=row, column=col)
                        setcol3.value = thisitem.shift_b_c or ''
                        colorfill = PatternFill(start_color=thisitem.bg_color_cell,
                                                end_color=thisitem.bg_color_cell, fill_type='solid')
                        setcol3.fill = colorfill

                    # for  Decimal or Number  Value Shift B
                    if thisitem.shift_b > 0 and thisitem.jobrouting_id:
                        setcol3 = worksheet.cell(row=row, column=col)
                        setcol3.value = thisitem.shift_b or ''
                        colorfill = PatternFill(start_color=thisitem.jobrouting_id.colour,
                                                end_color=thisitem.jobrouting_id.colour, fill_type='solid')

                        setcol3.fill = colorfill
                        # Setting value of comment if WELDER is Overloaded in Shift B
                        if thisitem.jobrouting_id.name == 'WELDING':
                            if commWELDING[col]:
                                setcol3.comment = commWELDING[col]
                        # END of Setting value of comment if WELDER is Overloaded in Shift B

                        # Setting value of comment if GRINDING is Overloaded in Shift B
                        if thisitem.jobrouting_id.name == 'GRINDING':
                            if commGRINDING[col]:
                                setcol3.comment = commGRINDING[col]
                        # END of Setting value of comment if GRINDING is Overloaded in Shift B

                        # Setting value of comment if Gouging is Overloaded in Shift B
                        if thisitem.jobrouting_id.name == 'Gouging':
                            if commGOUGING[col]:
                                setcol3.comment = commGOUGING[col]
                        # END of Setting value of comment if Gouging is Overloaded in Shift B

                    col = col + 1

                    # for  String  Value Shift C
                    if thisitem.shift_c_c != '':
                        setcol4 = worksheet.cell(row=row, column=col)
                        setcol4.value = thisitem.shift_c_c or ''
                        colorfill = PatternFill(start_color=thisitem.bg_color_cell,
                                                end_color=thisitem.bg_color_cell, fill_type='solid')
                        setcol4.fill = colorfill

                    # for  Decimal or Number  Value Shift C
                    if thisitem.shift_c > 0 and thisitem.jobrouting_id:
                        setcol4 = worksheet.cell(row=row, column=col)
                        setcol4.value = thisitem.shift_c
                        colorfill = PatternFill(start_color=thisitem.jobrouting_id.colour,
                                                end_color=thisitem.jobrouting_id.colour, fill_type='solid')

                        setcol4.fill = colorfill
                        # Setting value of comment if WELDER is Overloaded in Shift C
                        if thisitem.jobrouting_id.name == 'WELDING':
                            if commWELDING[col]:
                                setcol4.comment = commWELDING[col]
                        # END of Setting value of comment if WELDER is Overloaded in Shift C

                        # Setting value of comment if GRINDING is Overloaded in Shift B
                        if thisitem.jobrouting_id.name == 'GRINDING':
                            if commGRINDING[col]:
                                setcol4.comment = commGRINDING[col]
                        # END of Setting value of comment if GRINDING is Overloaded in Shift C

                        # Setting value of comment if Gouging is Overloaded in Shift C
                        if thisitem.jobrouting_id.name == 'Gouging':
                            if commGOUGING[col]:
                                setcol4.comment = commGOUGING[col]
                        # END of Setting value of comment if Gouging is Overloaded in Shift C

            row = row + 2
        # End of Section to generate Plan portion of the sheet
        # Start of Generation of Actual portion of sheet
        wb.active = 0
        worksheet = wb.active
        # Generating Actual Portion

        itemheader_obj = self.env['finishfetplanmodule.itemplanheadertable']
        itemheader_ids = itemheader_obj.search(
            ['&', ('wo_srno', '=', self.wo_srno), ('plan_date', '=', self.plan_date)])
        row = 27
        for thisitems_id in itemheader_ids:
            # Thse are already set while writing Plan portion. Row 15 is merged with Row 14 so write not possible
            # setcol1 = worksheet.cell(row=row, column=1)
            # setcol1.value = thisitems_id.name or ''
            # setcol1 = worksheet.cell(row=row, column=2)
            # setcol1.value = thisitems_id.wo_srno or ''

            for thisitem in thisitems_id.actualitemplan_id:
                if thisitem.date >= self.plan_date:
                    datediff = (thisitem.date - self.plan_date)
                    col = ((datediff.days + 1) * 3) + 1
                    dateStr = str(thisitem.date.day) + "/" + str(thisitem.date.month)
                    setdate = worksheet.cell(row=24, column=col)
                    setdate.value = dateStr

                    # for  String  Value Shift A
                    if thisitem.shift_a_c != '':
                        setcol2 = worksheet.cell(row=row, column=col)
                        setcol2.value = thisitem.shift_a_c or ''
                    # for  Decimal or Number  Value Shift A

                    if thisitem.shift_a > 0:
                        setcol2 = worksheet.cell(row=row, column=col)
                        setcol2.value = thisitem.shift_a or ''
                        colorfill = PatternFill(start_color=thisitem.jobrouting_id.colour,
                                                end_color=thisitem.jobrouting_id.colour, fill_type='solid')

                        setcol2.fill = colorfill
                    col = col + 1
                    # for  String  Value Shift B
                    if thisitem.shift_b_c != '':
                        setcol3 = worksheet.cell(row=row, column=col)
                        setcol3.value = thisitem.shift_b_c or ''

                    # for  Decimal or Number  Value Shift B
                    if thisitem.shift_b > 0:
                        setcol3 = worksheet.cell(row=row, column=col)
                        setcol3.value = thisitem.shift_b or ''
                        colorfill = PatternFill(start_color=thisitem.jobrouting_id.colour,
                                                end_color=thisitem.jobrouting_id.colour, fill_type='solid')

                        setcol3.fill = colorfill
                    col = col + 1
                    # for  String  Value Shift C
                    if thisitem.shift_c_c != '':
                        setcol4 = worksheet.cell(row=row, column=col)
                        setcol4.value = thisitem.shift_c_c or ''

                    # for  Decimal or Number  Value Shift C
                    if thisitem.shift_c > 0:
                        setcol4 = worksheet.cell(row=row, column=col)
                        setcol4.value = thisitem.shift_c
                        colorfill = PatternFill(start_color=thisitem.jobrouting_id.colour,
                                                end_color=thisitem.jobrouting_id.colour, fill_type='solid')

                        setcol4.fill = colorfill
            row = row + 2
        wb.active = 0
        worksheet = wb.active
        fp = io.BytesIO()
        wb.save(fp)
        out = base64.encodestring(fp.getvalue())
        # self.download_file = out
        view_empreport_id = self.env['view.empreport'].create(
            {'name': reportname, 'file_name': filename, 'datas_fname': out})
        return {
            'res_id': view_empreport_id.id,
            'name': 'Spent Report',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'view.empreport',
            'view_id': False,
            'type': 'ir.actions.act_window',
        }

    class view_empreport(models.TransientModel):
        _name = 'view.empreport'
        _rec_name = 'datas_fname'
        name = fields.Char('Report Name', size=256)
        file_name = fields.Char('File Name', size=256)
        datas_fname = fields.Binary('Report', size=256, default='Download Generated Plan.xlsx')
